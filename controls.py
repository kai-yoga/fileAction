import os
import time
# import win32com.client as win32
# import zipfile
import uuid
import vars
import xlrd
import openpyxl as opl
import datetime

class fileInfo:
    def __init__(self, fileName):
        self.fileName = fileName
        self._defaultPath=self.setDefaultPath()#defaultPath()
    def setDefaultPath(self):
        '''默认保存路径'''
        if not os.path.exists(vars.defaultPath):
            os.mkdir(vars.defaultPath)
        return vars.defaultPath
    def isExists(self):
        path,name=os.path.split(self.fileName)
        if not name:
            sign=0
        else:
            if not path:
                path=self._defaultPath
            sign=os.path.exists(os.path.join(path,name))
        # print(sign)
        return sign

    # def getFileInfo(self):
    #     '''文件的创建时间'''
    #     if self.isExists():
    #         creatTime = time.strftime('%Y-%m-%d', time.gmtime(os.path.getctime(self.fileName)))
    #         return {'createTime': creatTime}

    # def removeFile(self):
    #     '''删除文件'''
    #     if self.isExists():
    #         os.remove(self.fileName)
    #         return 1

    # def xlsToXlsx(self, mode='new'):
    #     '''从xls文件转换为xlsx文件'''
    #     if self.isExists() and self.fileName.lower().strip().endswith('xls'):
    #         excel = win32.gencache.EnsureDispatch('Excel.Application')
    #         wb = excel.Workbooks.Open(self.fileName)
    #         if os.path.exists(self.fileName + 'x'):
    #             if mode.lower().strip() == 'new':
    #                 os.remove(self.fileName + 'x')
    #                 wb.SaveAs(self.fileName+'x', FileFormat=51)
    #                 wb.Close()
    #                 excel.Application.Quit()
    #         else:
    #             wb.SaveAs(self.fileName + 'x', FileFormat=51)
    #             wb.Close()
    #             excel.Application.Quit()
    #     else:
    #         pass
    # def unzip(self,mode='new'):
    #     '''如果是压缩文件，则解压该文件'''
    #     if self.isExists() and zipfile.is_zipfile(self.fileName):
    #         zip=zipfile.ZipFile(self.fileName,'r')
    #         ziplist=zip.namelist()
    #         for z in ziplist:
    #             if os.path.exists(os.path.join(os.getcwd(),z)):
    #                 if mode=='new':
    #                     os.remove(os.path.join(os.getcwd(),z))
    #                     zip.extract(z)
    #             else:
    #                 zip.extract(z)
    def expFilename(self,suffix='.xlsx'):
        '''处理文件名
        :param filename:如果不包含路径，则默认路径，如果不指定文件名，则只用默认文件名，如果两者都不指定，那么都是默认
        :param content: 写入xls的内容，为二级嵌套格式，如[[]],或[()]
        :return:filename
        '''
        base = os.path.split(self.fileName)
        if base[0] and base[1]:
            path = base[0]
            if not base[1].endswith(suffix):
                name = base[1] + suffix
            else:
                name = base[1]
        elif base[0] and not base[1]:
            path = base[0]
            name = str(uuid.uuid1()).replace('-', '') + suffix
        elif not base[0] and base[1]:
            path = self._defaultPath
            if not base[1].endswith(suffix):
                name = base[1] + suffix
            else:
                name = base[1]
        else:
            path = self._defaultPath
            name = str(uuid.uuid1()).replace('-', '') + suffix
        pathname = os.path.join(path, name)
        return pathname
    def expXlsx(self, content=[], mode='new',suffix='.xlsx',sheetName='Sheet'):
        '''将content写入到指定的SheetName中
    '''
        xlsxFileName=self.expFilename(suffix)
        if suffix=='.xlsx':
            if os.path.exists(xlsxFileName):
                if mode=='new':
                    xlsxFileName=xlsxFileName+str(uuid.uuid1()).replace('-', '') + suffix
                    wb=opl.Workbook()
                else:
                    try:
                        wb=opl.load_workbook(xlsxFileName)
                    except:
                        wb=opl.Workbook()
            else:
                wb=opl.Workbook()
            wslist=wb.sheetnames
            if sheetName not in wslist:
                if sheetName is None:
                    sheetName='Sheet1'
                wb.create_sheet(sheetName)
            ws=wb[sheetName]
            ws.title=sheetName
            for data in content:
                ws.append(data)
            wb.save(xlsxFileName)
        elif suffix=='.txt':
            if os.path.exists(xlsxFileName):
                if mode=='new':
                    xlsxFileName=xlsxFileName+str(uuid.uuid1()).replace('-', '') + suffix
                    wb=open(xlsxFileName,'w',encoding='utf-8')
                else:
                    wb=open(xlsxFileName,'a',encoding='utf-8')
            else:
                wb=open(xlsxFileName,'w',encoding='utf-8')
            wb.write(content)
            wb.close()
        return 1
    # def removeXlsxWs(self,sheetName='Sheet1'):
    #     '''self本身必须是xlsx格式,且存在'''
    #     if self.isExists() and self.fileName.endswith('.xlsx'):
    #         wb=opl.load_workbook(self.fileName)
    #         for name in wb.get_sheet_names:
    #             wb.remove(name)
    #         wb.create_sheet(sheetName)
    def get_cell_value(self,cell):
        '''获取单元格的值，并且进行格式化，日期转换为XXXX-XX-XX,整数或浮点数转换为字符串'''
        cell_value=''
        if cell.value:
            try:
                if cell.is_date:
                    pass
                    year,month,day=cell.value.year,cell.value.month,cell.value.day
                    cell_value = datetime.date(year=year, month=month, day=day).strftime('%Y-%m-%d')
                else:
                    cell_value=str(cell.value).replace("‘",'').replace("’",'')
            except:
                if cell.ctype==3:
                    # pass
                    year,month,day=xlrd.xldate_as_tuple(cell.value,0)[0],xlrd.xldate_as_tuple(cell.value,0)[1],xlrd.xldate_as_tuple(cell.value,0)[2]
                # cell_value=str(year)+'-'+str(month)+'-'+str(day)
                    cell_value=datetime.date(year=year,month=month,day=day).strftime('%Y-%m-%d')
                else:
                    cell_value=str(cell.value).replace("‘",'').replace("’",'')
        return cell_value

    def getFileContent(self,sheetName='sheet1',type='active',containTitle=False):
        '''如果给定sheetName,则返回给定的SheetName的内容，如果没有找到相应的sheetName，如给定type='active'则
        返回活动工作表的内容，如不给定active则返回为空；
        如果不给定sheetName则返回所有工作表的内容！
        '''
        if self.isExists():
            if self.fileName.endswith('.xlsx'):
                wb=opl.load_workbook(self.fileName)
                wslist=wb.sheetnames
                L=[]
                ws=None
                if containTitle:
                    min_row=1
                else:
                    min_row=2
                if sheetName:
                    for name in wslist:
                        if name.lower().find(sheetName)>=0:
                            ws=wb[name]
                        else:
                            ws=None
                        if ws:
                            L.append(('内容来自于文件名={},工作表={}'.format(self.fileName,name),' '))
                            for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                rowL = []
                                for cell in row:
                                    rowL.append(self.get_cell_value(cell))
                                L.append(tuple(rowL))
                    if type=='active':
                        if ws is None and len(L)==0:
                            ws=wb.active
                            for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                rowL = []
                                for cell in row:
                                    rowL.append(self.get_cell_value(cell))
                                L.append(tuple(rowL))
                else:
                    for name in wslist:
                        ws=wb[name]
                        L.append(('内容来自于文件名={},工作表={}'.format(self.fileName, name), ' '))
                        for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                            rowL = []
                            for cell in row:
                                rowL.append(self.get_cell_value(cell))
                            L.append(tuple(rowL))
                if not len(L):
                    L=[('filename={}不存在指定的sheet={},且没有要求要返回active的工作表'.format(self.fileName,sheetName),'')]
                return L
            elif self.fileName.endswith('.xls'):
                # pass
                L=[]
                wb=xlrd.open_workbook(self.fileName)
                for ws in wb.sheets():
                    L.append(('内容来自于文件名={},工作表={}'.format(self.fileName, ws.name), ' '))
                    rows=ws.nrows
                    cols=ws.ncols
                    for row in range(rows):
                        rowL=[]
                        for col in range(cols):
                            cell=ws.cell(row,col)
                            rowL.append(self.get_cell_value(cell))
                        if len(rowL)>0:
                            L.append(tuple(rowL))
                return L
            else:
                return [('格式不受支持','格式不受支持')]

def pathCommon(path,type='1'):
    '''返回根目录下的子目录与文件集合{'dirs':resdirs,'files':resfiles}'''
    resdirs,resfiles=[],[]
    if path:
        if type=='1':
            for root,dirs,files in os.walk(path):
                for file in files:
                    resfiles.append(os.path.join(root,file))
                for dir in dirs:
                    resdirs.append(os.path.join(root,dir))
                    # pathCommon(os.path.join(path,dir),resdirs,resfiles)
        else:
            for root,dirs,files in os.walk(path):
                for file in files:
                    resfiles.append(os.path.join(path,file))
                break
            resdirs=path
        return {'dirs': resdirs, 'files': resfiles}


